"""
WiFi Usage Tracker

This script tracks the usage of devices connected to a WiFi router. It captures ARP, IP, and TCP packets
to extract information about device usage, including start time, total duration, and applications accessed.
The tracking data is stored in an Excel file, and the device status is printed after packet sniffing.

Author: Omar Arias
"""

import os
from scapy.all import sniff, ARP, IP, TCP
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
import pandas as pd
import threading
import sys
import time
import curses


class WiFiTracker:
    def __init__(self, excel_file='wifi_usage.xlsx'):
        self.excel_file = excel_file
        self.devices = {}
        self.start_time = None
        self.mydev_df = pd.DataFrame(columns=['MAC Addr', 'Device Name', 'Tipo', 'Descripción', 'Owner'])

        if os.path.exists(self.excel_file):
            self.load_existing_data()
        else:
            self.create_excel_file()
        self.get_device_list()

    def create_excel_file(self):
        return 0   # // Remover
        wb = Workbook()
        ws = wb.active
        ws.append(['Device', 'Start Time', 'End Time', 'Duration (s)', 'Applications'])
        wb.save(self.excel_file)

    def load_existing_data(self):
        return 0   # // Remover
        wb = load_workbook(self.excel_file)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            device, start_time, _, _, _ = row
            self.devices[device] = {'start_time': start_time, 'total_duration': 0, 'applications': set()}

        start_times = [data['start_time'] for data in self.devices.values()]
        self.start_time = min(start_times, default=None)

    def update_excel_file(self):
        return 0   # // Remover
        wb = load_workbook(self.excel_file)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, max_col=5, max_row=ws.max_row):
            for cell in row:
                cell.value = None

        for i, (device, data) in enumerate(self.devices.items(), start=2):
            ws.cell(row=i, column=1, value=device)
            ws.cell(row=i, column=2, value=data['start_time'])
            ws.cell(row=i, column=3, value=data['start_time'] + timedelta(seconds=data['total_duration']))
            ws.cell(row=i, column=4, value=data['total_duration'])
            ws.cell(row=i, column=5, value=', '.join(data['applications']))

        wb.save(self.excel_file)

    # Import devices info list in local WiFi
    def get_device_list(self):
        devices_file = 'ariher_devices.xlsx'
        
        if os.path.exists(devices_file):
            # Read the 'Devices' sheet from the Excel file into a DataFrame
            self.mydev_df = pd.read_excel(devices_file, sheet_name='Devices', usecols=['MAC Addr', 'Device Name', 'Tipo', 'Descripción', 'Owner'])
        else:
            # If the file doesn't exist, initialize an empty DataFrame
            self.mydev_df = pd.DataFrame(columns=['MAC Addr', 'Device Name', 'Tipo', 'Descripción', 'Owner'])

    # Returns a list with the device info, the elements:
    # 0 - Device Name
    # 1 - Tipo
    # 2 - Descripción
    # 3 - Owner
    def get_device_info(self, mac_address):
        # Check if the MAC address is in the DataFrame
        if mac_address in self.mydev_df['MAC Addr'].values:
            # Extract device info based on the MAC address
            device_info = self.mydev_df[self.mydev_df['MAC Addr'] == mac_address].iloc[0, 1:].tolist()
            return device_info
        else:
            # If MAC address not found, return None
            return None

    def print_connected_devices(self):
        print("\nConnected Devices:")
        for mac_address, data in self.devices.items():
            device_info = self.get_device_info(mac_address)

            if device_info is not None:
                device_name, tipo, descripcion, owner = device_info
                print(f"MAC Address: {mac_address}, Device Name: {device_name}, Tipo: {tipo}, Descripción: {descripcion}, Owner: {owner}")
                # print(f"Start Time: {data['start_time']}, Total Duration: {data['total_duration']} seconds, Applications: {', '.join(data['applications'])}")
                print(f"Start Time: {data['start_time']}, Total Duration: {int(data['total_duration'])} seconds")
                print ('Applications: ')
                print (*data['applications'], sep="\n")
                print("-" * 50)
            else:
                print(f"MAC Address: {mac_address}")
                # print(f"Start Time: {data['start_time']}, Total Duration: {int(data['total_duration'])} seconds, Applications: {', '.join(data['applications'])}")
                print(f"Start Time: {data['start_time']}, Total Duration: {int(data['total_duration'])} seconds")
                print ('Applications: ')
                print (*data['applications'], sep="\n")
                print("-" * 50)
                print('\n')
                
    '''                
    def packet_handler(self, pkt):
        if pkt.haslayer(ARP):
            device_mac = pkt[ARP].hwsrc
            current_time = datetime.now()

            print('packet_handler on device: ', device_mac)
            if device_mac not in self.devices:
                self.devices[device_mac] = {'start_time': current_time, 'total_duration': 0, 'applications': set()}
            else:
                previous_time = self.devices[device_mac]['start_time']
                duration = (current_time - previous_time).total_seconds()
                self.devices[device_mac]['total_duration'] += duration
                self.devices[device_mac]['start_time'] = current_time

        if pkt.haslayer(IP) and pkt.haslayer(TCP):
            device_mac = pkt.src
            destination_ip = pkt[IP].dst

            if device_mac in self.devices:
                self.devices[device_mac]['applications'].add(destination_ip)
    '''
    
    def top_devices(self, refresh_interval=5):
        def refresh_loop(stdscr):
            curses.curs_set(0)
            stdscr.clear()

            while True:
                self.update_devices_info()
                self.print_top_devices(stdscr)
                stdscr.refresh()
                time.sleep(refresh_interval)

        curses.wrapper(refresh_loop)

    def print_top_devices(self, stdscr):
        stdscr.clear()
        stdscr.addstr(0, 0, "[Press Q to quit]\n")
        stdscr.addstr(1, 0, "Connected Devices:")
        stdscr.addstr(2, 0, "{:<18} {:<20} {:<18} {:<25} {:<15} {:<5}".format('MAC Address', 'Device Name', 'Type', 'Description', 'Owner', 'Port 443'))
        stdscr.addstr(3, 0, "=" * 100)

        row = 4
        for mac_address, data in self.devices.items():
            device_info = self.get_device_info(mac_address)

            if device_info is not None:
                device_name, device_type, description, owner = device_info
                port_443_indicator = '443' if '443' in data['applications'] else ''
                stdscr.addstr(row, 0, "{:<18} {:<20} {:<18} {:<25} {:<15} {:<5}".format(mac_address, device_name, device_type, description, owner, port_443_indicator))
            else:
                stdscr.addstr(row, 0, "{:<18} {:<20} {:<18} {:<25} {:<15} {:<5}".format(mac_address, 'Unknown', 'N/A', 'N/A', 'N/A', ''))
            
            row += 1
            
                
    def update_devices_info(self):
        # Function to update device information (similar to your start_tracking method)
        sniff(prn=self.packet_handler, store=0, timeout=5, iface='en2')
        self.update_excel_file()
        
    def packet_handler(self, pkt):
        if pkt.haslayer(ARP):
            device_mac = pkt[ARP].hwsrc
            current_time = datetime.now()

            print('.', end='')
            # print('packet_handler on device: ', device_mac)
            if device_mac not in self.devices:
                self.devices[device_mac] = {'start_time': current_time, 'total_duration': 0, 'applications': set()}
            else:
                previous_time = self.devices[device_mac]['start_time']
                duration = (current_time - previous_time).total_seconds()
                self.devices[device_mac]['total_duration'] += duration
                self.devices[device_mac]['start_time'] = current_time

        if pkt.haslayer(IP) and pkt.haslayer(TCP):
            device_mac = pkt.src
            destination_ip = pkt[IP].dst
            destination_port = pkt[TCP].dport

            if device_mac in self.devices:
                self.devices[device_mac]['applications'].add(f"Port {destination_port}")

                # Map destination ports to applications (add more as needed)
                application_mapping = {
                    80: 'HTTP',
                    443: 'HTTPS (WS?)',
                    21: 'FTP',
                    22: 'SSH',
                    25: 'SMTP',
                    110: 'POP3',
                    143: 'IMAP',
                    53: 'DNS',
                    139: 'SMB',
                    445: 'SMB',
                    3389: 'RDP',
                }

                if destination_port in application_mapping:
                    application_name = application_mapping[destination_port]
                    self.devices[device_mac]['applications'].add(application_name)


    def start_tracking(self, duration=60):
        # sniff(prn=self.packet_handler, store=0, timeout=duration)
        print ('sniffing...')
        sniff(prn=self.packet_handler, store=0, timeout=duration, iface='en2')
        self.update_excel_file()

if __name__ == "__main__":
    wifi_tracker = WiFiTracker()
    wifi_tracker.start_tracking(duration=60)
    # wifi_tracker.print_connected_devices()
    wifi_tracker.top_devices(refresh_interval=5)